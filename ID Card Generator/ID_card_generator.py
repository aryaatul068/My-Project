import random
import datetime
import qrcode
from PIL import Image, ImageDraw, ImageFont
import os
import openpyxl  # Read the excel shett


os.system("Title: ID CARD Generator ")

d_date = datetime.datetime.now()
reg_format_date = d_date.strftime("  %d-%m-%Y\t\t\t\t\t ID CARD Generator\t\t\t\t\t  %I:%M:%S %p")
print(
    '+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
print(reg_format_date)
print(
    '+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')

wb = openpyxl.load_workbook("excelsheet1.xlsx")  # Or also you Give The Path of the  excel File

sheet = wb.sheetnames
sh1 = wb['Sheet1']
row = sh1.max_row        # Get the number of rows
column = sh1.max_column    # Get the Number of column

# print(row, column)
for i in range(2,row + 1):
    # for j in range(1, column + 1):
        # print(sh1.cell(i,1).value)
    image = Image.new('RGB', (1500, 900), (50, 200, 200))
    draw = ImageDraw.Draw(image)

# font = ImageFont.truetype('arial.ttf', size=12)


# Starting position of the message
    (x, y) = (330, 60)
    company = "Guru Ghasidas Vishwavidyalaya"
    color = 'rgb(0, 0, 0)'           # black color font
    font = ImageFont.truetype('arial.ttf', size=60)
    draw.text((x, y), company, fill=color, font=font)

#Adding an unique id number. You can manually take it from user
    (x, y) = (1150, 300)
    idno = random.randint(10000000, 90000000)
    message = str('ID: ' + str(idno))
    color = 'rgb(0, 0, 0)'           # black color font
    font = ImageFont.truetype('arial.ttf', size=40)
    draw.text((x, y), message, fill=color, font=font)

# For the Name
    (x, y) = (315, 260)
    name = sh1.cell(i,1).value

    fname = str('Name \t\t:' + str(name))
    color = 'rgb(0, 0, 0)'                    # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fname, fill=color, font=font)

# For the Father's Name
    (x, y) = (315, 300)
    faname = sh1.cell(i,2).value

    fname = str("Father's Name  : " + str(faname))
    color = 'rgb(0, 0, 0)'               # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fname, fill=color, font=font)

# For the Mother's Name
    (x, y) = (315, 340)
    mname = sh1.cell(i,3).value

    fname = str("Mother's Name   : " + str(mname))
    color = 'rgb(0, 0, 0)'                     # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fname, fill=color, font=font)

# For the Class
    (x, y) = (315, 380)
    class_ = sh1.cell(i,4).value

    fname = str('Class          : ' + str(class_))
    color = 'rgb(0, 0, 0)'                # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fname, fill=color, font=font)

# For the gender
    (x, y) = (315, 420)
    gender = sh1.cell(i,5).value
    fgender = str('Gender : ' + str(gender))
    color = 'rgb(0, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fgender, fill=color, font=font)
# For the Age
    (x, y) = (615, 420)
    age = sh1.cell(i,6).value
    fage = str('Age : ' + str(age))
    color = 'rgb(0, 0, 0)'               # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fage, fill=color, font=font)

# For the DOB
    (x, y) = (315, 460)
    dob = sh1.cell(i,7).value
    fdob = str('Date of Birth  : ' + str(dob))
    color = 'rgb(0, 0, 0)'                 # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fdob, fill=color, font=font)

# For the Blood Group
    (x, y) = (315, 500)
    blood_group= sh1.cell(i,8).value
    flood_group = str('Blood Group    : ' + str(blood_group))
    color = 'rgb(255, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), flood_group, fill=color, font=font)

# For the Mob No
    (x, y) = (315, 540)
    No = sh1.cell(i,9).value
    
    fNo = str('Mobile Number  : ' + str(No))
    color = 'rgb(0, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fNo, fill=color, font=font)

# For the Email Id.
# (x, y) = (400, 650)
# mail = int(input('Enter Your Email: '))

# fNo = str('Email: ' + str(mail))
# color = 'rgb(0, 0, 0)'            # black color
# draw.text((x, y), fNo, fill=color, font=font)

# For the Parent's Mob No
# (x, y) = (90, 280)
# pno = int(input("Enter Your Parent's Mobile Number: "))

# fNo = str("Parent's Number: " + str(pno))
# color = 'rgb(0, 0, 0)'                 # black color font
# draw.text((x, y), fNo, fill=color, font=font)

                # For the Address
# For Village/Colony and Post office Name
    (x, y) = (315, 580)
    villOrColPost = sh1.cell(i,10).value
    fvillOrColPost = str('Address :- ' + str(villOrColPost))
    color = 'rgb(200, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fvillOrColPost, fill=color, font=font)
        
# For District , State and Pincode
    (x, y) = (315, 620)
    distStatPin = sh1.cell(i,11).value
    fdistStatPin = str(str(distStatPin))
    color = 'rgb(0, 250, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fdistStatPin, fill=color, font=font)
        
# For student signature 
    (x, y) = (50, 760)
    fstud_sig = str("  Student \nSignature")
    color = 'rgb(0, 0, 40'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fstud_sig, fill=color, font=font)
# For student photo
    (x, y) = (70, 450)
    fstud_sig = str("Student\n Ptoto")
    color = 'rgb(0, 160, 0)'              # black color font
    draw.text((x, y), fstud_sig, fill=color, font=font)

# ID card Generate ---->> Date and Time 
    (x, y) = (370, 865)
    fstud_sig = d_date.strftime("ID Card Generate date and time    %d-%m-%Y  and   %I:%M:%S %p")
    color = 'rgb(0, 0, 0)'              # black color font
    draw.text((x, y), fstud_sig, fill=color, font=font)


# save the edited image
        
    image.save(str(name) + '.png')
    QR = qrcode.make('Company Name : '+ str(company) +'\nName : ' + str(name) +'\nID Number : ' + str(idno)+"\nFather's Name : " + str(faname) + "\nMother's Name : "+ str(mname) +'\nClass : '+ str(class_)+'\nGender : '+ str(gender) + "\nAge : " + str(age)+'\nDate of Birth : '+ str(dob) +'\nBlood Group : '+ str(blood_group) +'\nMobile No. : '+ str(No)  + '\nAddress : ' + str(villOrColPost) +  "\n" + str(distStatPin) , box_size = 4, border = 2, version = 12 )  # this info. is added in QR code
    QR.save(str(idno) + '.bmp')


    
    ID = Image.open(name + '.png')
    QR = Image.open(str(idno) + '.bmp')
        

# Logo of GGV
    img = Image.open("ggv.png") 
    resize_img = img.resize((300, 240)) # Resize the logo 
    ID.paste(resize_img, (0,0))
    
# parallel To x-axis ======>>>>>> Set
    x_axis_img = Image.open("red.jpg")
    resize_x_axis_image = x_axis_img.resize((5, 900))
    ID.paste(resize_x_axis_image, (300, 0))  # Draw line on the ID card        
# parallel To y-axis ========>>>>>>> set
    y_axis_img = Image.open("red.jpg")
    resize_y_axis_image = y_axis_img.resize((1500, 5))
    ID.paste(resize_y_axis_image, (0, 240)) # Draw line on the ID card

# This line Separate the image and signature ======>>>>>> Set
    y_axis_img = Image.open("red.jpg")
    resize_y_axis_image = y_axis_img.resize((300, 5))
    ID.paste(resize_y_axis_image, (0, 700)) # Draw line on the ID card

# Photo for ID Card Holder >>>>>>>>>Working

    # dp_img = Image.open("photo.jpg")
    # resize_dp_img = dp_img.resize((300,455))
    # ID.paste(resize_dp_img, (0, 245)) 

# QR for ID card
    ID.paste(QR, (1150, 380))
    ID.save(name + '.png')

print(('\n\n\nYour ID Card Successfully created in a PNG file '))