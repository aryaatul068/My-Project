#ID CARD GENERATOR SYSTEM
import datetime
import qrcode
from PIL import Image, ImageDraw, ImageFont
import os
import openpyxl                                             # Read the excel sheet ===>> intall : pip install openpyxl

from openpyxl import load_workbook                          # Image read from Excel sheet
from openpyxl_image_loader import SheetImageLoader          # Image read from Excel sheet

os.system("Title: ID Card Generator System ")

d_date = datetime.datetime.now()
reg_format_date = d_date.strftime("  %d-%m-%Y\t\t\t\t\t ID Card Generator System\t\t\t\t\t  %I:%M:%S %p")
print(
    '+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
print(reg_format_date)
print(
    '+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')

wb = openpyxl.load_workbook("excelsheet1.xlsx")             # Or also you Give The Path of the  excel File
sheet = wb.sheetnames
sh1 = wb['Sheet1']
row = sh1.max_row                                           # Get the number of rows
column = sh1.max_column                                     # Get the Number of column

image_loader = SheetImageLoader(sh1)

    # print(row, column)
for i in range(2,row+1):
    # for j in range(1, column + 1):
        # print(sh1.cell(i,1).value)
    image = Image.new('RGB', (1500, 900), (50, 200, 200))
    draw = ImageDraw.Draw(image)

# font = ImageFont.truetype('arial.ttf', size=12)


# Starting position of the message
    (x, y) = (330, 60)
    company = '''Guru Ghasidas Vishwavidyalaya
    Koni, Bilaspur, Chattisgarh'''
    color = 'rgb(0, 0, 0)'           # black color font
    font = ImageFont.truetype('arial.ttf', size=60)
  
    draw.text((x, y), company, fill=color, font=font)


#Adding an unique id number. You can manually take it from user
    (x, y) = (1000, 260)
    # print(str('Email ID'))
    idno = hash(sh1.cell(i,6).value)
    message = str('ID: ' + str(idno))
    color = 'rgb(0, 0, 0)'           # black color font
    font = ImageFont.truetype('arial.ttf', size=40)
    draw.text((x, y), message, fill=color, font=font)


# For the Name
    (x, y) = (315, 260)
    # name = sh1.cell(i,1).value
    fname = str('Name')
    color = 'rgb(0, 0, 0)'           # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fname, fill=color, font=font)


# Aline the all information
    (x, y) = (540, 260)
    name = sh1.cell(i,1).value
    fname = str(": " + str(name))
    color = 'rgb(0, 0, 0)'           # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fname, fill=color, font=font)


# For the Father's Name
    (x, y) = (315, 300)
    # faname = sh1.cell(i,2).value
    fname = str("Father's Name")
    color = 'rgb(0, 0, 0)'           # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fname, fill=color, font=font)


# Aline the all information
    (x, y) = (540, 300)
    faname = sh1.cell(i,2).value
    fname = str(": " + str(faname))
    color = 'rgb(0, 0, 0)'           # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fname, fill=color, font=font)


# For the Mother's Name
    (x, y) = (315, 340)
    # mname = sh1.cell(i,3).value
    fname = str("Mother's Name")
    color = 'rgb(0, 0, 0)'           # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fname, fill=color, font=font)


# Aline the all information
    (x, y) = (540, 340)
    mname = sh1.cell(i,3).value
    fname = str(": " + str(mname))
    color = 'rgb(0, 0, 0)'           # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fname, fill=color, font=font)


# For the Class
    (x, y) = (315, 380)
    # class_ = sh1.cell(i,4).value
    fname = str('Class')
    color = 'rgb(0, 0, 0)'           # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fname, fill=color, font=font)


# Aline the all information
    (x, y) = (540, 380)
    class_ = sh1.cell(i,4).value
    fname = str(': ' + str(class_))
    color = 'rgb(0, 0, 0)'                # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fname, fill=color, font=font)


# For the gender
    (x, y) = (315, 420)
    # gender = sh1.cell(i,5).value
    fgender = str('Gender')
    color = 'rgb(0, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fgender, fill=color, font=font)


# Aline the all information
    (x, y) = (540, 420)
    gender = sh1.cell(i,5).value
    fgender = str(': ' + str(gender))
    color = 'rgb(0, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fgender, fill=color, font=font)

# For the Email Id
    (x, y) = (315, 460)
    # emailid = int(sh1.cell(i,6).value)
    femailid = str('Email ID')
    color = 'rgb(0, 0, 0)'               # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), femailid, fill=color, font=font)




# Aline the all information
    (x, y) = (540, 460)
    emailid = sh1.cell(i,6).value
    femailid = str(': ' + str(emailid))
    color = 'rgb(0, 0, 0)'               # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), femailid, fill=color, font=font)


# For the DOB
    (x, y) = (315, 500)
    # dob = sh1.cell(i,7).value
    fdob = str('Date of Birth')
    color = 'rgb(0, 0, 0)'                 # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fdob, fill=color, font=font)


# Aline the all information
    (x, y) = (540, 500)
    dob = sh1.cell(i,7).value
    fdob = str(': ' + str(dob))
    color = 'rgb(0, 0, 0)'                 # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fdob, fill=color, font=font)


# For the Blood Group
    (x, y) = (315, 540)
    # blood_group= sh1.cell(i,8).value
    flood_group = str('Blood Group')
    color = 'rgb(0, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), flood_group, fill=color, font=font)


# Aline the all information
    (x, y) = (540, 540)
    blood_group= sh1.cell(i,8).value
    flood_group = str(': ' + str(blood_group))
    color = 'rgb(0, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), flood_group, fill=color, font=font)


# For the Mob No
    (x, y) = (315, 580)
    # No = int(sh1.cell(i,9).value)
    fNo = str('Mobile Number')
    color = 'rgb(0, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fNo, fill=color, font=font)


# Aline the all information
    (x, y) = (540, 580)
    No = int(sh1.cell(i,9).value)
    fNo = str(': +91 ' + str(No))
    color = 'rgb(0, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fNo, fill=color, font=font)


# For the Address
    # For Village/Colony and Post office Name
    (x, y) = (315, 620)
    # villOrColPost = sh1.cell(i,10).value
    fvillOrColPost = str('Address')
    color = 'rgb(0, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fvillOrColPost, fill=color, font=font)


# Aline the all information
    (x, y) = (540, 620)
    villOrColPost = sh1.cell(i,10).value
    fvillOrColPost = str(': ' + str(villOrColPost) + ",")
    color = 'rgb(0, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fvillOrColPost, fill=color, font=font)
        

# For District , State and Pincode
    (x, y) = (555, 660)
    distStatPin = sh1.cell(i,11).value
    fdistStatPin = str(str(distStatPin))
    color = 'rgb(0, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fdistStatPin, fill=color, font=font)
        

# For student signature 
    (x, y) = (10, 710)
    fstud_sig = str("Signature of Student")
    color = 'rgb(0, 0, 0)'              # black color font
    font = ImageFont.truetype('arial.ttf', size=30)
    draw.text((x, y), fstud_sig, fill=color, font=font)


# For student photo
    (x, y) = (70, 450)
    fstud_sig = str("Student\n Ptoto")
    color = 'rgb(0, 0, 0)'              # black color font
    draw.text((x, y), fstud_sig, fill=color, font=font)


# ID card Generate ---->> Date and Time 
    (x, y) = (650, 865)
    fstud_sig = d_date.strftime("    ID Card Generated date : %d-%m-%Y ")
    color = 'rgb(0, 0, 0)'              # black color font
    draw.text((x, y), fstud_sig, fill=color, font=font)


# save the edited image
    
    image.save(str(name) + '.png')
    QR = qrcode.make('Company Name : '+ str(company) +'\nName : ' + str(name) +'\nID Number : ' + str(idno)+"\nFather's Name : " +
     str(faname) + "\nMother's Name : "+ str(mname) +'\nClass : '+ str(class_)+'\nGender : '+ str(gender) + "\nAge : " + 
     str(emailid)+'\nDate of Birth : '+ str(dob) +'\nBlood Group : '+ str(blood_group) +'\nMobile No. : '+ str(No)  + '\nAddress : ' + 
     str(villOrColPost) +  "\n" + str(distStatPin) , box_size = 4, border = 2, version = 12 )  # this info. is added in QR code
    QR.save(str(idno) + '.bmp')

    
    ID = Image.open( name + '.png')
    QR = Image.open(str(idno) + '.bmp')
        

# Logo of GGV
    img = Image.open("ggv.png") 
    resize_img = img.resize((300, 240)) # Resize the logo 
    ID.paste(resize_img, (0,0))
    

# parallel To x-axis ======>>>>>> Set
    x_axis_img = Image.open("red.jpg")
    resize_x_axis_image = x_axis_img.resize((5, 900))
    ID.paste(resize_x_axis_image, (300, 0))  # Draw line on the ID card        


# parallel To y-axis ========>>>>>>> set
    y_axis_img = Image.open("red.jpg")
    resize_y_axis_image = y_axis_img.resize((1500, 5))
    ID.paste(resize_y_axis_image, (0, 240)) # Draw line on the ID card


# This line Separate the image and signature ======>>>>>> Set
    y_axis_img = Image.open("red.jpg")
    resize_y_axis_image = y_axis_img.resize((300, 5))
    ID.paste(resize_y_axis_image, (0, 700)) # Draw line on the ID card


# Photo of ID Card Holder >>>>>>>>>Working
    
    photo_path = f"L{i}"
    dp_img = image_loader.get(photo_path)
    resize_dp_img = dp_img.resize((300,455))
    ID.paste(resize_dp_img, (0, 245)) 
    

# Signature of ID Card Holder >>>>>>>>>Working
    
    sig_path = f"M{i}"
    sig_img = image_loader.get(sig_path)
    resize_sig_img = sig_img.resize((300,150))
    ID.paste(resize_sig_img, (0, 750)) 


# QR for ID card
    ID.paste(QR, (1150, 320))
    ID.save(name + '.png')

print(('\n\n \t\t\t Your ID Card Successfully generated in a PNG file \n\n'))